import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def read_excel_file(file_path):
    """读取Excel文件"""
    try:
        df = pd.read_excel(file_path, sheet_name=0, header=0)
        print(f"成功读取文件，共 {len(df)} 行数据")
        return df
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return None


def merge_person_cells_in_excel(input_file, output_file):
    """
    直接合并相同人员的单元格，不进行任何数值计算
    合并：序号、姓名、身份证号码、工资总额 这四列
    """

    print("正在读取Excel文件...")
    # 先读取数据，了解人员分布
    df = pd.read_excel(input_file, sheet_name=0)

    # 直接保存原文件内容到新文件
    print("正在保存副本文件...")
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="人员-公司工资小结", index=False)

    # 重新打开文件进行合并单元格操作
    print("正在合并相同人员的单元格...")
    wb = load_workbook(output_file)
    ws = wb.active

    # 获取数据范围
    max_row = ws.max_row
    max_col = ws.max_column

    # 列索引（从1开始）
    COL_SEQ = 1  # A列 - 序号
    COL_NAME = 2  # B列 - 姓名
    COL_ID = 3  # C列 - 身份证号码
    COL_TOTAL_SALARY = 7  # G列 - 工资总额

    print(f"将合并以下列: 序号(A列), 姓名(B列), 身份证号码(C列), 工资总额(G列)")

    # 从第2行开始（跳过表头）
    row = 2
    merge_count = 0
    while row <= max_row:
        current_name = ws.cell(row=row, column=COL_NAME).value
        current_id = ws.cell(row=row, column=COL_ID).value
        current_salary = ws.cell(row=row, column=COL_TOTAL_SALARY).value

        if current_name is None:  # 跳过空行
            row += 1
            continue

        # 找到相同人员的起始和结束行
        start_row = row
        end_row = row

        # 向下查找相同姓名和身份证号的人员
        for next_row in range(row + 1, max_row + 1):
            next_name = ws.cell(row=next_row, column=COL_NAME).value
            next_id = ws.cell(row=next_row, column=COL_ID).value

            # 如果姓名和身份证号都相同，则属于同一个人
            if next_name == current_name and next_id == current_id:
                end_row = next_row
            else:
                break

        # 如果有多行相同人员，合并单元格
        if end_row > start_row:
            merge_count += 1
            print(f"合并人员: {current_name}, 行 {start_row} 到 {end_row}")

            # 合并序号列（A列）
            ws.merge_cells(
                start_row=start_row,
                start_column=COL_SEQ,
                end_row=end_row,
                end_column=COL_SEQ,
            )

            # 合并姓名列（B列）
            ws.merge_cells(
                start_row=start_row,
                start_column=COL_NAME,
                end_row=end_row,
                end_column=COL_NAME,
            )

            # 合并身份证号码列（C列）
            ws.merge_cells(
                start_row=start_row,
                start_column=COL_ID,
                end_row=end_row,
                end_column=COL_ID,
            )

            # 合并工资总额列（G列）
            ws.merge_cells(
                start_row=start_row,
                start_column=COL_TOTAL_SALARY,
                end_row=end_row,
                end_column=COL_TOTAL_SALARY,
            )

            # 设置合并后单元格的垂直对齐方式为居中
            for col in [COL_SEQ, COL_NAME, COL_ID, COL_TOTAL_SALARY]:
                cell = ws.cell(row=start_row, column=col)
                cell.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )

        # 跳到下一组人员
        row = end_row + 1

    # 调整列宽
    print("正在调整列宽...")
    for column in range(1, max_col + 1):
        max_length = 0
        column_letter = get_column_letter(column)

        # 检查表头长度
        header = ws.cell(row=1, column=column).value
        if header:
            max_length = len(str(header))

        # 检查数据长度
        for row_num in range(2, min(max_row, 100) + 1):  # 检查前100行
            cell_value = ws.cell(row=row_num, column=column).value
            if cell_value:
                cell_length = len(str(cell_value))
                if cell_length > max_length:
                    max_length = cell_length

        # 设置列宽（限制最大宽度）
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # 特别调整身份证号码列宽度
    ws.column_dimensions["C"].width = 22  # 身份证号码列
    ws.column_dimensions["G"].width = 15  # 工资总额列

    # 冻结首行
    ws.freeze_panes = "A2"

    # 保存文件
    wb.save(output_file)

    print(f"\n处理完成！结果已保存到: {output_file}")

    # 输出统计信息
    unique_names = df["姓名"].nunique()
    total_rows = len(df)
    print(f"\n统计信息:")
    print(f"原始数据行数: {total_rows}")
    print(f"实际人员数量: {unique_names}")
    print(f"合并后节省行数: {total_rows - unique_names}")
    print(f"合并的人员组数: {merge_count}")


def main():
    # 文件路径
    input_file = "step1.xlsx"
    output_file = "step2_merge.xlsx"

    # 检查文件是否存在
    if not Path(input_file).exists():
        print(f"错误: 找不到文件 {input_file}")
        return

    # 执行合并操作
    merge_person_cells_in_excel(input_file, output_file)

    print("\n提示: 打开Excel文件后，你可以看到相同人员的以下列已经合并为一个单元格:")
    print("  - 序号 (A列)")
    print("  - 姓名 (B列)")
    print("  - 身份证号码 (C列)")
    print("  - 工资总额 (G列)")


if __name__ == "__main__":
    main()
