import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def read_excel_file(file_path):
    """读取Excel文件"""
    try:
        df = pd.read_excel(file_path, sheet_name=0, header=0)
        print(f"成功读取文件，共 {len(df)} 行数据")
        return df
    except Exception as e:
        print(f"读取文件时出错: {e}")
        return None


def preprocess_data(df):
    """数据预处理：填充空值，处理合并单元格"""
    df_processed = df.copy()

    # 向前填充序号、姓名、身份证号等关键信息（处理合并单元格）
    columns_to_fill = [
        "序号",
        "姓名",
        "身份证号码",
        "实际工作单位名称",
        "实际工作单位税号",
        "用工性质",
        "工资总额",
    ]

    for col in columns_to_fill:
        if col in df_processed.columns:
            df_processed[col] = df_processed[col].fillna(method="ffill")

    # 删除没有工资或工作单位的行
    df_processed = df_processed.dropna(subset=["工资", "实际工作单位名称"], how="all")

    # 确保工资列为数值类型
    if "工资" in df_processed.columns:
        df_processed["工资"] = pd.to_numeric(df_processed["工资"], errors="coerce")
        df_processed = df_processed.dropna(subset=["工资"])
        df_processed = df_processed[df_processed["工资"] > 0]

    print(f"预处理后，共 {len(df_processed)} 行有效数据")
    return df_processed


def summarize_by_person_company(df):
    """
    按人员和公司汇总工资
    将同一个人员在同一个公司的多个月份工资汇总成一条记录
    """

    print("正在按人员和公司汇总工资...")

    # 按人员和公司分组汇总
    person_company_summary = (
        df.groupby(
            [
                "序号",
                "姓名",
                "身份证号码",
                "实际工作单位名称",
                "实际工作单位税号",
                "用工性质",
                "工资总额",  # 年度总工资
            ]
        )
        .agg(
            {
                "工资": ["sum", "count"],  # 在该公司的工资总额、工作月数
                # "工作时间": lambda x: sorted(list(x)),  # 列出工作月份（排序）
            }
        )
        .reset_index()
    )

    # 重命名列
    person_company_summary.columns = [
        "序号",
        "姓名",
        "身份证号码",
        "工作单位名称",
        "单位税号",
        "用工性质",
        "年度工资总额",
        "公司工资小计",
        "工作月数",
        "工作月份列表",
    ]

    # 计算平均月薪（公司工资小计 / 工作月数）
    person_company_summary["平均月薪"] = (
        person_company_summary["公司工资小计"] / person_company_summary["工作月数"]
    ).round(2)

    # 添加工作月份范围
    person_company_summary["工作月份范围"] = person_company_summary[
        "工作月份列表"
    ].apply(lambda x: f"{x[0]} 至 {x[-1]}" if len(x) > 1 else x[0])

    # 按序号排序
    person_company_summary = person_company_summary.sort_values("序号").reset_index(
        drop=True
    )

    # 重新排列列顺序
    cols = [
        "序号",
        "姓名",
        "身份证号码",
        "工作单位名称",
        "单位税号",
        "用工性质",
        "年度工资总额",
        "公司工资小计",
        "工作月数",
        "平均月薪",
        "工作月份列表",
        "工作月份范围",
    ]
    person_company_summary = person_company_summary[cols]

    print(f"汇总完成，共 {len(person_company_summary)} 条人员-公司记录")

    # 显示汇总示例
    print("\n=== 汇总示例（丁学娟）===")
    example = person_company_summary[person_company_summary["姓名"] == "丁学娟"]
    if len(example) > 0:
        print(
            example[
                [
                    "姓名",
                    "工作单位名称",
                    "年度工资总额",
                    "公司工资小计",
                    "工作月数",
                    "工作月份列表",
                ]
            ]
        )

    return person_company_summary


def create_company_summary(df):
    """按公司汇总"""

    company_summary = (
        df.groupby(["工作单位名称", "单位税号"])
        .agg(
            {
                "姓名": lambda x: list(x),  # 列出所有人员
                "公司工资小计": "sum",  # 公司支付的总工资
                "工作月数": "sum",  # 总工作月数
            }
        )
        .reset_index()
    )

    company_summary["人员数量"] = company_summary["姓名"].apply(len)
    company_summary.columns = [
        "工作单位名称",
        "单位税号",
        "人员列表",
        "公司总工资",
        "总工作月数",
        "人员数量",
    ]

    # 按公司总工资降序排序
    company_summary = company_summary.sort_values(
        "公司总工资", ascending=False
    ).reset_index(drop=True)

    return company_summary


def save_results(person_company_summary, company_summary, output_file):
    """保存结果到Excel"""

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        # 人员-公司工资汇总
        person_company_summary.to_excel(
            writer, sheet_name="人员-公司工资汇总", index=False
        )

        # 公司汇总
        company_summary.to_excel(writer, sheet_name="公司汇总", index=False)

        # 统计信息
        stats_data = {
            "统计项": [
                "总记录数",
                "涉及人员数",
                "涉及单位数",
                "总工资额",
                "平均每人工资",
            ],
            "数值": [
                len(person_company_summary),
                person_company_summary["姓名"].nunique(),
                len(company_summary),
                person_company_summary["年度工资总额"].sum(),
                person_company_summary["年度工资总额"].mean(),
            ],
        }
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name="统计信息", index=False)

    print(f"\n结果已保存到: {output_file}")


def apply_excel_styling(output_file):
    """应用Excel样式"""

    wb = load_workbook(output_file)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # 设置标题行加粗
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)

        # 调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = "A2"

    wb.save(output_file)
    print("Excel样式应用完成")


def main():
    # 文件路径
    input_file = "step1.xlsx"
    output_file = "step1_merge.xlsx"

    # 检查文件是否存在
    if not Path(input_file).exists():
        print(f"错误: 找不到文件 {input_file}")
        return

    # 读取数据
    print("正在读取Excel文件...")
    df = read_excel_file(input_file)
    if df is None:
        return

    # 数据预处理
    df_processed = preprocess_data(df)

    # 按人员和公司汇总
    person_company_summary = summarize_by_person_company(df_processed)

    # 按公司汇总
    company_summary = create_company_summary(person_company_summary)

    # 保存结果
    save_results(person_company_summary, company_summary, output_file)

    # 应用样式
    apply_excel_styling(output_file)

    print("\n处理完成！")
    print(f"原始数据行数: {len(df)}")
    print(f"预处理后行数: {len(df_processed)}")
    print(f"人员-公司汇总记录数: {len(person_company_summary)}")
    print(f"涉及人员数: {person_company_summary['姓名'].nunique()}")
    print(f"涉及单位数: {len(company_summary)}")


if __name__ == "__main__":
    main()
